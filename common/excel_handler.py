#coding = uft-8
import openpyxl
import os

base_path = os.path.dirname(os.getcwd())

class HandlerExcel:
    def __init__(self,path):
        self.path = path


    def open_excel(self):
        open_excel = openpyxl.load_workbook(base_path + self.path)
        return open_excel

    def get_sheet(self,index=None):
        sheet_name = self.open_excel().sheetnames
        if index == None:
            index = 0
        return self.open_excel()[sheet_name[index]]

    def get_excel(self):
        ws = self.get_sheet()
        data = ws.cell(1,1).value
        print(data)

if __name__ == "__main__":
    Handler = HandlerExcel('/case/case.xlsx')
    Handler.get_excel()





